import openpyxl

path = "C:\\Users\\Aaryan\\Downloads\\shreyapythonfiles\\Book1.xlsx"

workbook = openpyxl.load_workbook(path)
#sheet = workbook["Sheet1"]
sheet = workbook.active #active fn gives the title of the default first worksheet.
rows = sheet.max_row
cols = sheet.max_column
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="       ")
    print()
    