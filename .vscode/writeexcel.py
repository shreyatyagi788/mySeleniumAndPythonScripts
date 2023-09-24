import openpyxl
# # writing into empty excel file having sheet name 'Data'.

# path = "C:\\Users\\Aaryan\\Downloads\\shreyapythonfiles\\emptydata.xlsx"
# workbook = openpyxl.load_workbook(path)
# #sheet = workbook['Data']
# sheet = workbook.active #if we have only one sheet then we can use 'active'

# #this will write same data in all the cells.
# for i in range(1,6):
#     for j in range(1,4):
#         sheet.cell(i,j).value = "welcome"
# workbook.save(path)

#writing multiple data.
file = "C:\\Users\\Aaryan\\Downloads\\shreyapythonfiles\\testdata.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "shreya"
sheet.cell(1,3).value = "engineer"

sheet.cell(2,1).value = 124
sheet.cell(2,2).value = "mini"
sheet.cell(2,3).value = "developer"

workbook.save(file)

